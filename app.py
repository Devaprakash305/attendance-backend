from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import pandas as pd
import openpyxl

app = Flask(__name__)
CORS(app)

TOTAL_STUDENTS = 64
STUDENT_FILE = "students.xlsx"

# Load student data from students.xlsx
try:
    df = pd.read_excel(STUDENT_FILE)
    roll_to_name = {}
    for _, row in df.iterrows():
        # Assuming the Excel has columns for roll number and name
        # Adjust column names based on your actual Excel structure
        roll_col = 'Roll No' if 'Roll No' in df.columns else df.columns[0]
        name_col = 'Name' if 'Name' in df.columns else df.columns[1]
        roll_to_name[int(row[roll_col])] = row[name_col]
except Exception as e:
    print(f"Error loading student data: {e}")
    roll_to_name = {}

@app.route("/attendance", methods=["POST"])
def attendance():
    data = request.json

    date = data.get("date")
    hour = data.get("hour")
    absent = data.get("absent", "")
    od = data.get("od", "")

    if not date or not hour:
        return jsonify({"error": "Date and Hour are required"}), 400

    absent_rolls = sorted(set(
        int(x) for x in absent.split(",") if x.strip().isdigit()
    ))
    od_rolls = sorted(set(
        int(x) for x in od.split(",") if x.strip().isdigit()
    ))

    # Check overlap
    common = set(absent_rolls) & set(od_rolls)
    if common:
        return jsonify({
            "warning": f"Roll numbers present in BOTH Absentees and OD: {', '.join(map(str, common))}"
        })

    present = TOTAL_STUDENTS - len(absent_rolls)
    percentage = round((present / TOTAL_STUDENTS) * 100)

    absentees = [
        f"{roll_to_name[r]} ({r})"
        for r in absent_rolls if r in roll_to_name
    ]

    ods = [
        f"{roll_to_name[r]} ({r})"
        for r in od_rolls if r in roll_to_name
    ]

    # --- Attendance.xlsx update logic ---
    ATTENDANCE_FILE = "Attendance.xlsx"
    from openpyxl.utils import get_column_letter

    wb = openpyxl.load_workbook(ATTENDANCE_FILE)
    ws = wb.active

    # Find reg no column (assume first col), and header row
    header = [cell.value for cell in ws[1]]
    reg_col_idx = 1  # 1-based index

    # Check if date column exists, else add
    date_col = None
    for idx, col in enumerate(header, 1):
        if str(col).strip() == date:
            date_col = idx
            break
    if not date_col:
        date_col = len(header) + 1
        ws.cell(row=1, column=date_col, value=date)

    # Mark attendance for each student
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        reg_no = row[reg_col_idx-1].value
        if reg_no is None:
            continue
        try:
            reg_no_int = int(str(reg_no).strip())
        except Exception:
            continue
        if reg_no_int in absent_rolls:
            mark = "A"
        elif reg_no_int in od_rolls:
            mark = "OD"
        else:
            mark = "P"
        ws.cell(row=row[0].row, column=date_col, value=mark)

    wb.save(ATTENDANCE_FILE)

    # --- End Attendance.xlsx update logic ---

    result = f"""Good morning sir,\nDate : {date}\nHour: {hour}\nII YEAR - A\nB.Tech IT  : {present}/64\n--------------------------------\nPercentage : {percentage}%\n\n *Absentees List*\n"""
    for i, a in enumerate(absentees, 1):
        result += f"{i}. {a}\n"
    result += "\nOD\n"
    for i, o in enumerate(ods, 1):
        result += f"{i}. {o}\n"
    result += "\nThank you sir"

    return jsonify({"output": result})



# Route to download Attendance.xlsx
@app.route("/download-attendance", methods=["GET"])
def download_attendance():
    return send_file("Attendance.xlsx", as_attachment=True)

if __name__ == "__main__":
    app.run()
